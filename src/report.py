from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import matplotlib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .storage import Punch


matplotlib.use("Agg")
import matplotlib.pyplot as plt


REPORT_DIR = Path("reports")
PNG_DIR = REPORT_DIR / "png"
INVALID_SHEET_CHARS = set("\\/*?:[]")
INVALID_FILE_CHARS = set('<>:"/\\|?*')
WEEKDAY_LABELS = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
DEFAULT_DAILY_TARGET = timedelta(hours=8)
WORK_MODE_STANDARD = "STANDARD"
WORK_MODE_HOME = "HOME"
WORK_MODE_FIELD = "CAMPO"
WORK_MODE_LABELS = {
    WORK_MODE_STANDARD: "",
    WORK_MODE_HOME: "Home Office",
    WORK_MODE_FIELD: "Campo",
}
WORK_MODE_COLORS = {
    WORK_MODE_HOME: ("DBEAFE", "1E3A8A"),
    WORK_MODE_FIELD: ("FEF3C7", "92400E"),
}


@dataclass(slots=True)
class DailySummary:
    day: date
    entry: datetime | None
    lunch_out: datetime | None
    return_in: datetime | None
    day_exit: datetime | None
    total_worked: timedelta
    target: timedelta
    work_mode: str
    has_events: bool
    pending: bool


@dataclass(slots=True)
class UserMonthlySummary:
    user_id: int
    user_name: str
    rows: list[DailySummary]
    month_total: timedelta
    month_balance: timedelta


def _month_range(target_year: int, target_month: int) -> tuple[date, date]:
    start = date(target_year, target_month, 1)
    if target_month == 12:
        end = date(target_year + 1, 1, 1)
    else:
        end = date(target_year, target_month + 1, 1)
    return start, end


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if ch in INVALID_SHEET_CHARS else ch for ch in name).strip()
    return (cleaned or "Usuario")[:31]


def _safe_file_component(name: str) -> str:
    cleaned = "".join("_" if ch in INVALID_FILE_CHARS else ch for ch in name).strip()
    cleaned = "_".join(cleaned.split())
    return (cleaned or "usuario")[:64]


def _summarize_day(
    events: list[Punch],
) -> tuple[timedelta, bool, datetime | None, datetime | None, datetime | None, datetime | None]:
    total = timedelta()
    pending = False

    entry: datetime | None = None
    lunch_out: datetime | None = None
    return_in: datetime | None = None
    day_exit: datetime | None = None
    stack_in: datetime | None = None

    for event in events:
        action = event.action

        if event.action == "ENTRY":
            if entry is None:
                entry = event.ts_utc
            else:
                pending = True
        elif event.action == "LUNCH_OUT":
            if lunch_out is None:
                lunch_out = event.ts_utc
            else:
                pending = True
        elif event.action == "RETURN":
            if return_in is None:
                return_in = event.ts_utc
            else:
                pending = True
        elif event.action == "EXIT":
            if day_exit is None:
                day_exit = event.ts_utc
            else:
                pending = True
        elif event.action == "IN":
            if stack_in is None:
                stack_in = event.ts_utc
                if entry is None:
                    entry = event.ts_utc
                elif lunch_out is not None and return_in is None:
                    return_in = event.ts_utc
            else:
                pending = True
        elif event.action == "OUT":
            if entry is not None and lunch_out is None:
                lunch_out = event.ts_utc
            elif return_in is not None and day_exit is None:
                day_exit = event.ts_utc
        else:
            pending = True

        if action in {"ENTRY", "RETURN", "IN"}:
            if stack_in is None:
                stack_in = event.ts_utc
            else:
                pending = True
        elif action in {"LUNCH_OUT", "EXIT", "OUT"}:
            if stack_in is not None:
                total += event.ts_utc - stack_in
                stack_in = None
            else:
                pending = True

    if stack_in is not None:
        pending = True

    if lunch_out is not None and entry is None:
        pending = True
    if return_in is not None and lunch_out is None and entry is not None:
        pending = True
    if day_exit is not None and return_in is None:
        pending = True

    return total, pending, entry, lunch_out, return_in, day_exit


def _format_day_with_weekday(day: date) -> str:
    return f"{day.strftime('%d/%m/%Y')} ({WEEKDAY_LABELS[day.weekday()]})"


def _daily_balance(total_worked: timedelta, has_events: bool, target: timedelta) -> tuple[str, str]:
    if not has_events:
        return "", "empty"

    diff = total_worked - target
    return _format_signed_duration_hhmm(diff)


def _format_signed_duration_hhmm(diff: timedelta) -> tuple[str, str]:
    total_minutes = int(abs(diff.total_seconds()) // 60)
    hours, minutes = divmod(total_minutes, 60)

    if diff > timedelta():
        return f"+ {hours:02d}:{minutes:02d}", "positive"
    if diff < timedelta():
        return f"-{hours:02d}:{minutes:02d}", "negative"
    return "00:00", "neutral"


def _format_duration_hhmm(total: timedelta) -> str:
    total_minutes = int(total.total_seconds() // 60)
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours:02d}:{minutes:02d}"


def _resolve_daily_target(
    user_id: int,
    day: date,
    default_daily_target: timedelta,
    daily_target_overrides_by_user: dict[int, dict[date, timedelta]] | None,
) -> timedelta:
    if daily_target_overrides_by_user is None:
        return default_daily_target
    user_overrides = daily_target_overrides_by_user.get(user_id)
    if user_overrides is None:
        return default_daily_target
    return user_overrides.get(day, default_daily_target)


def _resolve_work_mode(
    user_id: int,
    day: date,
    work_modes_by_user: dict[int, dict[date, str]] | None,
) -> str:
    if work_modes_by_user is None:
        return WORK_MODE_STANDARD
    user_modes = work_modes_by_user.get(user_id)
    if user_modes is None:
        return WORK_MODE_STANDARD
    return user_modes.get(day, WORK_MODE_STANDARD)


def _build_user_summaries(
    punches: list[Punch],
    tz_name: str,
    target_year: int,
    target_month: int,
    default_daily_target: timedelta = DEFAULT_DAILY_TARGET,
    daily_target_overrides_by_user: dict[int, dict[date, timedelta]] | None = None,
    work_modes_by_user: dict[int, dict[date, str]] | None = None,
) -> list[UserMonthlySummary]:
    tz = ZoneInfo(tz_name)
    month_start, month_end = _month_range(target_year, target_month)

    grouped: dict[int, dict[date, list[Punch]]] = defaultdict(lambda: defaultdict(list))
    latest_name_by_user: dict[int, str] = {}
    latest_ts_by_user: dict[int, datetime] = {}

    for punch in punches:
        local_ts = punch.ts_utc.astimezone(tz)
        local_day = local_ts.date()
        if local_day < month_start or local_day >= month_end:
            continue
        grouped[punch.user_id][local_day].append(
            Punch(
                id=punch.id,
                chat_id=punch.chat_id,
                user_id=punch.user_id,
                user_name=punch.user_name,
                action=punch.action,
                ts_utc=local_ts,
            )
        )
        prev_ts = latest_ts_by_user.get(punch.user_id)
        if prev_ts is None or local_ts >= prev_ts:
            latest_ts_by_user[punch.user_id] = local_ts
            latest_name_by_user[punch.user_id] = punch.user_name

    summaries: list[UserMonthlySummary] = []

    for user_id, days in sorted(grouped.items(), key=lambda item: latest_name_by_user.get(item[0], "").lower()):
        user_name = latest_name_by_user.get(user_id, str(user_id))
        rows: list[DailySummary] = []
        month_total = timedelta()

        cursor = month_start
        while cursor < month_end:
            events = days.get(cursor, [])
            daily_target = _resolve_daily_target(
                user_id,
                cursor,
                default_daily_target,
                daily_target_overrides_by_user,
            )
            work_mode = _resolve_work_mode(user_id, cursor, work_modes_by_user)
            total, pending, entry, lunch_out, return_in, day_exit = _summarize_day(events)
            month_total += total
            rows.append(
                DailySummary(
                    day=cursor,
                    entry=entry,
                    lunch_out=lunch_out,
                    return_in=return_in,
                    day_exit=day_exit,
                    total_worked=total,
                    target=daily_target,
                    work_mode=work_mode,
                    has_events=bool(events),
                    pending=pending,
                )
            )
            cursor += timedelta(days=1)

        month_target = sum((row.target for row in rows if row.has_events), timedelta())
        month_balance = month_total - month_target

        summaries.append(
            UserMonthlySummary(
                user_id=user_id,
                user_name=user_name,
                rows=rows,
                month_total=month_total,
                month_balance=month_balance,
            )
        )

    return summaries


def build_month_report(
    punches: list[Punch],
    tz_name: str,
    target_year: int,
    target_month: int,
    default_daily_target: timedelta = DEFAULT_DAILY_TARGET,
    daily_target_overrides_by_user: dict[int, dict[date, timedelta]] | None = None,
    work_modes_by_user: dict[int, dict[date, str]] | None = None,
) -> Path:
    summaries = _build_user_summaries(
        punches,
        tz_name,
        target_year,
        target_month,
        default_daily_target=default_daily_target,
        daily_target_overrides_by_user=daily_target_overrides_by_user,
        work_modes_by_user=work_modes_by_user,
    )

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "Data (Dia)",
        "Entrada",
        "Saida Almoco",
        "Entrada 2",
        "Saida Final",
        "Horas Trabalhadas",
        "Saldo (8h)",
        "Modo",
        "Pendente",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for summary in summaries:
        ws = wb.create_sheet(title=_safe_sheet_name(summary.user_name))
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for row_idx, row in enumerate(summary.rows, start=2):
            balance_text, balance_kind = _daily_balance(row.total_worked, row.has_events, row.target)
            ws.append(
                [
                    _format_day_with_weekday(row.day),
                    row.entry.strftime("%H:%M") if row.entry else "",
                    row.lunch_out.strftime("%H:%M") if row.lunch_out else "",
                    row.return_in.strftime("%H:%M") if row.return_in else "",
                    row.day_exit.strftime("%H:%M") if row.day_exit else "",
                    _format_duration_hhmm(row.total_worked),
                    balance_text,
                    WORK_MODE_LABELS.get(row.work_mode, ""),
                    "SIM" if row.pending else "",
                ]
            )
            zebra_fill = PatternFill("solid", fgColor="F7FAFC" if row_idx % 2 == 0 else "FFFFFF")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = zebra_fill
                cell.alignment = center
            if balance_kind == "positive":
                ws.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor="DCFCE7")
                ws.cell(row=row_idx, column=7).font = Font(color="166534", bold=True)
            elif balance_kind == "negative":
                ws.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor="FECACA")
                ws.cell(row=row_idx, column=7).font = Font(color="991B1B", bold=True)
            if row.work_mode in WORK_MODE_COLORS:
                fill_color, font_color = WORK_MODE_COLORS[row.work_mode]
                ws.cell(row=row_idx, column=8).fill = PatternFill("solid", fgColor=fill_color)
                ws.cell(row=row_idx, column=8).font = Font(color=font_color, bold=True)
            if row.pending:
                ws.cell(row=row_idx, column=9).fill = PatternFill("solid", fgColor="FECACA")
                ws.cell(row=row_idx, column=9).font = Font(color="991B1B", bold=True)

        ws.append([])
        total_row = ws.max_row + 1
        month_balance_text, month_balance_kind = _format_signed_duration_hhmm(summary.month_balance)
        ws.append(["TOTAL MENSAL", "", "", "", "", _format_duration_hhmm(summary.month_total), month_balance_text, "", ""])
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="E2E8F0")
        ws.cell(row=total_row, column=6).fill = PatternFill("solid", fgColor="E2E8F0")
        ws.cell(row=total_row, column=7).fill = PatternFill("solid", fgColor="E2E8F0")
        ws.cell(row=total_row, column=1).alignment = center
        ws.cell(row=total_row, column=6).alignment = center
        ws.cell(row=total_row, column=7).alignment = center
        if month_balance_kind == "positive":
            ws.cell(row=total_row, column=7).font = Font(color="166534", bold=True)
        elif month_balance_kind == "negative":
            ws.cell(row=total_row, column=7).font = Font(color="991B1B", bold=True)

        ws.append([])
        ws.append(["LEGENDA", "", "", "", "", "", "", "", ""])
        legend_header_row = ws.max_row
        ws.cell(row=legend_header_row, column=1).font = Font(bold=True)
        ws.append(["HOME", "Home Office", "", "", "", "", "", "", ""])
        ws.append(["CAMPO", "Campo", "", "", "", "", "", "", ""])
        ws.cell(row=legend_header_row + 1, column=1).fill = PatternFill("solid", fgColor=WORK_MODE_COLORS[WORK_MODE_HOME][0])
        ws.cell(row=legend_header_row + 1, column=1).font = Font(color=WORK_MODE_COLORS[WORK_MODE_HOME][1], bold=True)
        ws.cell(row=legend_header_row + 2, column=1).fill = PatternFill("solid", fgColor=WORK_MODE_COLORS[WORK_MODE_FIELD][0])
        ws.cell(row=legend_header_row + 2, column=1).font = Font(color=WORK_MODE_COLORS[WORK_MODE_FIELD][1], bold=True)

        widths = [18, 10, 12, 10, 11, 16, 13, 14, 10]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Sem registros")
        ws.append(["Nao houve registros no periodo."])

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    output = REPORT_DIR / f"ponto_{target_year}_{target_month:02d}.xlsx"
    wb.save(output)
    return output


def _render_user_table_png(summary: UserMonthlySummary, target_year: int, target_month: int) -> Path:
    PNG_DIR.mkdir(parents=True, exist_ok=True)
    filename = (
        f"ponto_{target_year}_{target_month:02d}_"
        f"{summary.user_id}_{_safe_file_component(summary.user_name)}.png"
    )
    output = PNG_DIR / filename

    headers = [
        "Data (Dia)",
        "Entrada",
        "Saida Almoco",
        "Entrada 2",
        "Saida Final",
        "Horas Trabalhadas",
        "Saldo (8h)",
        "Modo",
        "Pendente",
    ]
    rows = [
        [
            _format_day_with_weekday(row.day),
            row.entry.strftime("%H:%M") if row.entry else "",
            row.lunch_out.strftime("%H:%M") if row.lunch_out else "",
            row.return_in.strftime("%H:%M") if row.return_in else "",
            row.day_exit.strftime("%H:%M") if row.day_exit else "",
            _format_duration_hhmm(row.total_worked),
            _daily_balance(row.total_worked, row.has_events, row.target)[0],
            WORK_MODE_LABELS.get(row.work_mode, ""),
            "SIM" if row.pending else "",
        ]
        for row in summary.rows
    ]
    month_balance_text, _month_balance_kind = _format_signed_duration_hhmm(summary.month_balance)
    rows.append(["TOTAL MENSAL", "", "", "", "", _format_duration_hhmm(summary.month_total), month_balance_text, "", ""])

    fig_height = max(4.5, len(rows) * 0.35)
    fig, ax = plt.subplots(figsize=(15.5, fig_height))
    fig.patch.set_facecolor("#F8FAFC")
    ax.set_facecolor("#F8FAFC")
    ax.axis("off")

    table = ax.table(
        cellText=rows,
        colLabels=headers,
        loc="center",
        cellLoc="center",
        colColours=["#1F4E78"] * len(headers),
        colWidths=[0.17, 0.08, 0.1, 0.08, 0.09, 0.14, 0.11, 0.12, 0.08],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.5)
    table.scale(1.0, 1.25)

    for (r, c), cell in table.get_celld().items():
        if r == 0:
            cell.set_text_props(color="white", weight="bold")
            cell.set_edgecolor("#D1D5DB")
        else:
            cell.set_edgecolor("#E5E7EB")
            cell.set_facecolor("#F8FAFC" if r % 2 == 0 else "#FFFFFF")
            if c == 6:
                if rows[r - 1][6].startswith("+ "):
                    cell.set_facecolor("#DCFCE7")
                    cell.set_text_props(color="#166534", weight="bold")
                elif rows[r - 1][6].startswith("-"):
                    cell.set_facecolor("#FECACA")
                    cell.set_text_props(color="#991B1B", weight="bold")
            if c == 8 and rows[r - 1][8] == "SIM":
                cell.set_facecolor("#FECACA")
                cell.set_text_props(color="#991B1B", weight="bold")
            if c == 7:
                if rows[r - 1][7] == "Home Office":
                    cell.set_facecolor("#DBEAFE")
                    cell.set_text_props(color="#1E3A8A", weight="bold")
                elif rows[r - 1][7] == "Campo":
                    cell.set_facecolor("#FEF3C7")
                    cell.set_text_props(color="#92400E", weight="bold")

    total_row = len(rows)
    for c in range(len(headers)):
        cell = table[(total_row, c)]
        cell.set_facecolor("#E2E8F0")
        if c in {0, 5, 6}:
            cell.set_text_props(weight="bold")
    if rows[total_row - 1][6].startswith("+ "):
        table[(total_row, 6)].set_text_props(color="#166534", weight="bold")
    elif rows[total_row - 1][6].startswith("-"):
        table[(total_row, 6)].set_text_props(color="#991B1B", weight="bold")

    plt.title(
        f"Ponto {target_month:02d}/{target_year} - {summary.user_name}",
        fontsize=13,
        fontweight="bold",
        pad=14,
        color="#0F172A",
    )
    plt.suptitle(
        "Resumo diario de jornada",
        fontsize=9,
        y=0.97,
        color="#475569",
    )
    fig.text(
        0.02,
        0.02,
        "HOME",
        fontsize=8,
        color="#1E3A8A",
        bbox={"facecolor": "#DBEAFE", "edgecolor": "#93C5FD", "boxstyle": "round,pad=0.25"},
    )
    fig.text(0.10, 0.02, "Home Office", fontsize=8, color="#334155")
    fig.text(
        0.26,
        0.02,
        "CAMPO",
        fontsize=8,
        color="#92400E",
        bbox={"facecolor": "#FEF3C7", "edgecolor": "#FDE68A", "boxstyle": "round,pad=0.25"},
    )
    fig.text(0.36, 0.02, "Campo", fontsize=8, color="#334155")
    plt.tight_layout(rect=(0, 0.04, 1, 0.95))
    fig.savefig(output, dpi=200, bbox_inches="tight")
    plt.close(fig)

    return output


def build_month_report_images(
    punches: list[Punch],
    tz_name: str,
    target_year: int,
    target_month: int,
    default_daily_target: timedelta = DEFAULT_DAILY_TARGET,
    daily_target_overrides_by_user: dict[int, dict[date, timedelta]] | None = None,
    work_modes_by_user: dict[int, dict[date, str]] | None = None,
) -> list[tuple[str, Path]]:
    summaries = _build_user_summaries(
        punches,
        tz_name,
        target_year,
        target_month,
        default_daily_target=default_daily_target,
        daily_target_overrides_by_user=daily_target_overrides_by_user,
        work_modes_by_user=work_modes_by_user,
    )

    outputs: list[tuple[str, Path]] = []
    for summary in summaries:
        outputs.append((summary.user_name, _render_user_table_png(summary, target_year, target_month)))

    if outputs:
        return outputs

    PNG_DIR.mkdir(parents=True, exist_ok=True)
    empty_file = PNG_DIR / f"ponto_{target_year}_{target_month:02d}_sem_registros.png"

    fig, ax = plt.subplots(figsize=(8, 2.5))
    ax.axis("off")
    ax.text(0.5, 0.5, "Nao houve registros no periodo.", ha="center", va="center", fontsize=12)
    plt.tight_layout()
    fig.savefig(empty_file, dpi=180, bbox_inches="tight")
    plt.close(fig)

    return [("Sem registros", empty_file)]
